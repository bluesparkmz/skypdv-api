from fastapi import APIRouter, Depends, HTTPException, status, Query, UploadFile, File, Request
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, or_
from typing import List, Optional
from datetime import datetime
import io

from database import get_db
from auth import get_current_user
from models import User, PDVStockMovement, MovementType, PDVSale, PDVSaleItem, PDVProduct, PDVInventory
import schemas
from controllers import controller
import openpyxl
from openpyxl.utils import get_column_letter
from whatsapp_service import send_whatsapp_file, send_whatsapp_text

# Criar router principal
router = APIRouter(
    prefix="/skypdv",
    tags=["skypdv"]
)

# ===================================================================
# Terminal Endpoints
# ===================================================================

@router.get("/terminal", response_model=schemas.PDVTerminal)
def get_my_terminal(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Obter o terminal PDV do usuário atual (requer setup)"""
    return controller.get_terminal_required(db, current_user.id)


@router.post("/terminal/setup", response_model=schemas.PDVTerminal)
def setup_my_terminal(
    terminal_data: schemas.PDVTerminalCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Fazer setup do PDV (criar terminal manualmente para lojas/farmácias)"""
    return controller.create_terminal_for_user(db, current_user.id, terminal_data)

@router.post("/terminal", response_model=schemas.PDVTerminal)
def create_my_terminal(
    terminal_data: schemas.PDVTerminalCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Alias compatível para criação de terminal (POST /terminal)."""
    return controller.create_terminal_for_user(db, current_user.id, terminal_data)

@router.put("/terminal", response_model=schemas.PDVTerminal)
def update_my_terminal(
    terminal_update: schemas.PDVTerminalUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Atualizar configurações do terminal"""
    terminal = controller.get_terminal_required(db, current_user.id)
    return controller.update_terminal(db, terminal.id, terminal_update, current_user.id)

# ===================================================================
# Terminal Users Management - Gestão de usuários do terminal
# ===================================================================

@router.get("/terminal/users", response_model=List[schemas.PDVTerminalUser])
def list_terminal_users(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Lista todos os usuários associados ao terminal"""
    terminal = controller.get_terminal_required(db, current_user.id)
    return controller.get_terminal_users(db, terminal.id, current_user.id)


@router.post("/terminal/users", response_model=schemas.PDVTerminalUser)
def add_terminal_user(
    user_data: schemas.PDVTerminalUserCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Adiciona um usuário ao terminal pelo email"""
    terminal = controller.get_terminal_required(db, current_user.id)
    return controller.add_terminal_user(db, terminal.id, user_data.email, user_data, current_user.id)


@router.put("/terminal/users/{terminal_user_id}", response_model=schemas.PDVTerminalUser)
def update_terminal_user(
    terminal_user_id: int,
    user_update: schemas.PDVTerminalUserUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Atualiza permissões de um usuário do terminal"""
    terminal = controller.get_terminal_required(db, current_user.id)
    return controller.update_terminal_user(db, terminal.id, terminal_user_id, user_update, current_user.id)


@router.delete("/terminal/users/{terminal_user_id}")
def remove_terminal_user(
    terminal_user_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Remove um usuário do terminal"""
    terminal = controller.get_terminal_required(db, current_user.id)
    controller.remove_terminal_user(db, terminal.id, terminal_user_id, current_user.id)
    return {"message": "User removed from terminal successfully"}

# ===================================================================
# Suppliers Endpoints
# ===================================================================

@router.get("/suppliers", response_model=List[schemas.PDVSupplier])
def list_suppliers(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Listar fornecedores conectados ao terminal"""
    terminal = controller.get_terminal_required(db, current_user.id)
    return controller.get_suppliers(db, terminal.id)

@router.post("/suppliers", response_model=schemas.PDVSupplier)
def add_supplier(
    supplier: schemas.PDVSupplierCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Adicionar novo fornecedor manual"""
    terminal = controller.get_terminal_required(db, current_user.id)
    return controller.create_supplier(db, supplier, terminal.id)

@router.post("/suppliers/connect/fastfood", response_model=schemas.PDVSupplier)
def connect_fastfood(
    data: schemas.ConnectFastFoodRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Conectar um restaurante FastFood como fornecedor"""
    terminal = controller.get_or_create_terminal(db, current_user.id)
    return controller.connect_fastfood_restaurant(db, terminal.id, data.restaurant_id, data.sync_products)

@router.post("/suppliers/{supplier_id}/sync", response_model=schemas.PDVSupplier)
def force_sync_supplier(
    supplier_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Sincronizar manualmente produtos de um fornecedor externo (FastFood)"""
    terminal = controller.get_terminal_required(db, current_user.id)
    return controller.sync_supplier(db, supplier_id, terminal.id)

@router.get("/suppliers/{supplier_id}/products", response_model=List[schemas.PDVProduct])
def list_supplier_products(
    supplier_id: int,
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Listar produtos de um fornecedor específico (ex: FastFood)"""
    terminal = controller.get_terminal_required(db, current_user.id)
    return controller.get_products(
        db,
        terminal.id,
        supplier_id=supplier_id,
        limit=limit,
        skip=skip
    )

@router.put("/suppliers/{supplier_id}", response_model=schemas.PDVSupplier)
def update_supplier(
    supplier_id: int,
    updates: schemas.PDVSupplierUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Atualizar fornecedor"""
    terminal = controller.get_terminal_required(db, current_user.id)
    return controller.update_supplier(db, supplier_id, updates, terminal.id)

@router.delete("/suppliers/{supplier_id}")
def delete_supplier(
    supplier_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Desativar fornecedor"""
    terminal = controller.get_terminal_required(db, current_user.id)
    return controller.delete_supplier(db, supplier_id, terminal.id)

# ===================================================================
# Products & Inventory Endpoints
# ===================================================================

@router.get("/products", response_model=List[schemas.PDVProduct])
def list_products(
    search: Optional[str] = None,
    category: Optional[str] = None,
    source_type: Optional[schemas.SourceTypeEnum] = None,
    is_fastfood: Optional[bool] = None,
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Listar produtos com filtros (incluindo is_fastfood para FastFood)"""
    terminal = controller.get_terminal_required(db, current_user.id)
    return controller.get_products(
        db, terminal.id, 
        search=search, 
        category=category, 
        source_type=source_type,
        is_fastfood=is_fastfood,
        limit=limit,
        skip=skip
    )


@router.get("/products/category-summary/today", response_model=schemas.PDVCategorySalesSummary)
def get_category_sales_summary_today(
    category: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    terminal = controller.get_terminal_required(db, current_user.id)
    return controller.get_category_sales_summary_today(db, terminal.id, category)


@router.get("/products/category-report", response_model=schemas.PDVCategorySalesReport)
def get_category_sales_report(
    category: str,
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    user_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    terminal = controller.get_terminal_required(db, current_user.id)
    if not controller.is_terminal_admin(db, terminal.id, current_user.id):
        user_id = current_user.id
    return controller.get_category_sales_report(
        db,
        terminal.id,
        category,
        start_date=start_date,
        end_date=end_date,
        user_id=user_id,
    )

@router.post("/products", response_model=schemas.PDVProduct)
def create_product(
    product: schemas.PDVProductCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Criar novo produto no PDV"""
    terminal = controller.get_terminal_required(db, current_user.id)
    controller.require_terminal_permission(db, terminal.id, current_user.id, "can_manage_products")
    return controller.create_product(db, product, terminal.id)


@router.get("/products/catalog", response_model=List[schemas.PDVProduct])
def list_shared_products(
    search: Optional[str] = None,
    category: Optional[str] = None,
    business_type: Optional[str] = Query(None, pattern="^(loja|restaurante)$"),
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    terminal = controller.get_terminal_required(db, current_user.id)
    return controller.get_shared_products(
        db,
        terminal.id,
        search=search,
        category=category,
        business_type=business_type,
        limit=limit,
        skip=skip,
    )


@router.post("/products/adopt", response_model=schemas.PDVProduct)
def adopt_shared_product(
    payload: schemas.PDVProductAdopt,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    terminal = controller.get_terminal_required(db, current_user.id)
    controller.require_terminal_permission(db, terminal.id, current_user.id, "can_manage_products")
    return controller.adopt_shared_product(
        db,
        terminal.id,
        payload.source_product_id,
        price=payload.price,
        cost_price=payload.cost_price,
        initial_stock=payload.initial_stock,
    )

@router.post("/products/upload-image")
async def upload_product_image(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Upload an image for a PDV product.
    Returns the URL of the uploaded image.
    """
    # Verify user has access to a terminal (is a valid PDV user)
    terminal = controller.get_terminal_required(db, current_user.id)
    controller.require_terminal_permission(db, terminal.id, current_user.id, "can_manage_products")
    
    url = await controller.upload_pdv_product_image(file)
    return {"url": url}


@router.post("/invoice-assets/upload")
async def upload_invoice_asset(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Upload logo or stamp image for invoice settings.
    """
    terminal = controller.get_terminal_required(db, current_user.id)
    controller.require_terminal_permission(db, terminal.id, current_user.id, "can_sell")

    url = await controller.upload_pdv_invoice_asset(file)
    return {"url": url}

@router.post("/products/search", response_model=List[schemas.PDVProduct])
def search_products(
    search: schemas.PDVProductSearch,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Busca avançada de produtos com múltiplos filtros (incluindo is_fastfood e supplier_id)"""
    terminal = controller.get_terminal_required(db, current_user.id)
    return controller.get_products(
        db,
        terminal.id,
        search=search.query,
        category=search.category,
        source_type=search.source_type.value if search.source_type else None,
        is_fastfood=search.is_fastfood,
        supplier_id=search.supplier_id,
        limit=search.limit,
        skip=search.skip
    )

@router.get("/categories", response_model=List[str])
def list_categories(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Listar todas as categorias de produtos"""
    terminal = controller.get_terminal_required(db, current_user.id)
    return controller.get_product_categories(db, terminal.id)

@router.get("/products/stats", response_model=schemas.PDVProductStats)
def get_product_stats(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Estatísticas de produtos (total, ativos, FastFood, locais, categorias)"""
    terminal = controller.get_terminal_required(db, current_user.id)
    return controller.get_product_stats(db, terminal.id)

@router.patch("/products/{product_id}", response_model=schemas.PDVProduct)
@router.put("/products/{product_id}", response_model=schemas.PDVProduct, include_in_schema=False)
def update_product(
    product_id: int,
    updates: schemas.PDVProductUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Atualizar produto"""
    terminal = controller.get_terminal_required(db, current_user.id)
    controller.require_terminal_permission(db, terminal.id, current_user.id, "can_manage_products")
    return controller.update_product(db, product_id, updates, terminal.id)

@router.put("/products/batch/fastfood", response_model=List[schemas.PDVProduct])
def batch_update_fastfood_flag(
    batch: schemas.PDVProductBatchFastFood,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Marcar/desmarcar produtos como FastFood em lote"""
    terminal = controller.get_terminal_required(db, current_user.id)
    controller.require_terminal_permission(db, terminal.id, current_user.id, "can_manage_products")
    return controller.batch_update_fastfood_flag(db, batch.product_ids, batch.is_fastfood, terminal.id)

@router.delete("/products/{product_id}")
def delete_product(
    product_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Desativar um produto (Marcar como inativo)"""
    terminal = controller.get_terminal_required(db, current_user.id)
    controller.require_terminal_permission(db, terminal.id, current_user.id, "can_manage_products")
    return controller.delete_product(db, product_id, terminal.id)

@router.get("/products/{product_id}/movements", response_model=List[schemas.PDVStockMovement])
def list_stock_movements(
    product_id: int,
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Histórico de movimentações de stock de um produto"""
    terminal = controller.get_terminal_required(db, current_user.id)
    return controller.get_stock_movements(db, terminal.id, product_id, skip, limit)

@router.get("/inventory/movements", response_model=List[schemas.PDVStockMovement])
def list_inventory_movements(
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Histórico global de movimentações de stock do terminal"""
    terminal = controller.get_terminal_required(db, current_user.id)
    return controller.get_stock_movements(db, terminal.id, None, skip, limit)

@router.post("/inventory/adjustment", response_model=schemas.PDVStockMovement)
def adjust_inventory(
    adjustment: schemas.StockAdjustment,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Ajustar estoque manual (entrada/saída/balanço)"""
    terminal = controller.get_terminal_required(db, current_user.id)
    controller.require_terminal_permission(db, terminal.id, current_user.id, "can_manage_stock")
    return controller.adjust_stock(db, adjustment, terminal.id, current_user.id)

@router.post("/inventory/transfer", response_model=schemas.PDVStockMovement)
def transfer_inventory(
    transfer: schemas.StockTransfer,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Transferir estoque entre localizações"""
    terminal = controller.get_terminal_required(db, current_user.id)
    controller.require_terminal_permission(db, terminal.id, current_user.id, "can_manage_stock")
    return controller.transfer_stock(db, transfer, terminal.id, current_user.id)

@router.get("/inventory", response_model=schemas.InventoryReport)
def get_inventory_report(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Relatório detalhado de inventário e stock baixo"""
    terminal = controller.get_terminal_required(db, current_user.id)
    return controller.get_inventory_report(db, terminal.id)

@router.put("/inventory/{product_id}", response_model=schemas.PDVInventory)
def update_inventory_config(
    product_id: int,
    updates: schemas.PDVInventoryUpdate,
    storage_location: str = "balcao",
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Atualizar configuracoes do inventario por local"""
    terminal = controller.get_terminal_required(db, current_user.id)
    controller.require_terminal_permission(db, terminal.id, current_user.id, "can_manage_stock")
    return controller.update_inventory_settings(db, product_id, terminal.id, storage_location, updates)

# ===================================================================
# Cash Register Endpoints
# ===================================================================

@router.get("/cash-register/current", response_model=Optional[schemas.PDVCashRegister])
def get_current_register(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Obter sessão do caixa atual"""
    terminal = controller.get_terminal_required(db, current_user.id)
    return controller.get_current_register(db, terminal.id, current_user.id)

@router.post("/cash-register/open", response_model=schemas.PDVCashRegister)
def open_register(
    data: schemas.PDVCashRegisterOpen,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Abrir o caixa"""
    terminal = controller.get_terminal_required(db, current_user.id)
    controller.require_terminal_permission(db, terminal.id, current_user.id, "can_open_cash_register")
    return controller.open_register(db, data, terminal.id, current_user.id)

@router.post("/cash-register/close", response_model=schemas.PDVCashRegister)
def close_register(
    data: schemas.PDVCashRegisterClose,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Fechar o caixa"""
    terminal = controller.get_terminal_required(db, current_user.id)
    controller.require_terminal_permission(db, terminal.id, current_user.id, "can_open_cash_register")
    return controller.close_register(db, data, terminal.id, current_user.id)


@router.get("/cash-register/history", response_model=List[schemas.PDVCashRegister])
def list_cash_registers(
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    user_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Histórico de caixas (apenas admins podem filtrar por usuário)."""
    terminal = controller.get_terminal_required(db, current_user.id)
    # Se não for admin, força usar apenas o próprio user_id
    if not controller.is_terminal_admin(db, terminal.id, current_user.id):
        user_id = current_user.id
    return controller.list_cash_registers(db, terminal.id, start_date, end_date, user_id)

# ===================================================================
# Sales Endpoints
# ===================================================================

@router.post("/sales", response_model=schemas.PDVSale)
def create_sale(
    sale: schemas.PDVSaleCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Registrar nova venda"""
    terminal = controller.get_terminal_required(db, current_user.id)
    controller.require_terminal_permission(db, terminal.id, current_user.id, "can_sell")
    return controller.create_sale(db, sale, terminal.id, current_user.id)

@router.get("/sales", response_model=List[schemas.PDVSale])
def list_sales(
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    source_type: Optional[str] = None,
    payment_method: Optional[str] = None,
    sale_type: Optional[str] = None,
    status: Optional[str] = "completed",
    skip: int = 0,
    limit: int = 50,
    user_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Listar histórico de vendas com filtros.
    - Caixas veem apenas suas próprias vendas
    - Admins veem todas as vendas do terminal
    - Se user_id for fornecido e usuário for admin, filtra por esse caixa específico
    Permite visualizar:
    - Por período de data
    - Por origem da venda (mantido por compatibilidade)
    - Por tipo de venda (sale_type: local, delivery, online)
    - Por método de pagamento
    """
    terminal = controller.get_terminal_required(db, current_user.id)
    if controller.is_terminal_admin(db, terminal.id, current_user.id):
        filter_user_id = user_id
    else:
        filter_user_id = current_user.id
    return controller.get_sales(
        db, terminal.id, 
        skip=skip, 
        limit=limit,
        start_date=start_date,
        end_date=end_date,
        source_type=source_type,
        payment_method=payment_method,
        sale_type=sale_type,
        status=status,
        user_id=filter_user_id
    )

# ===================================================================
# Invoice Endpoints (usam o mesmo modelo de venda)
# ===================================================================

@router.post("/invoices", response_model=schemas.PDVSale)
def create_invoice(
    sale: schemas.PDVSaleCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    terminal = controller.get_terminal_required(db, current_user.id)
    controller.require_terminal_permission(db, terminal.id, current_user.id, "can_sell")
    return controller.create_invoice(db, sale, terminal.id, current_user.id)

@router.get("/invoices", response_model=List[schemas.PDVSale])
def list_invoices(
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    payment_status: Optional[str] = None,
    skip: int = 0,
    limit: int = 50,
    user_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    terminal = controller.get_terminal_required(db, current_user.id)
    if controller.is_terminal_admin(db, terminal.id, current_user.id):
        filter_user_id = user_id
    else:
        filter_user_id = current_user.id
    status = None  # retorna todas para ver pendentes e pagas
    sales = controller.get_sales(
        db, terminal.id,
        skip=skip,
        limit=limit,
        start_date=start_date,
        end_date=end_date,
        status=status,
        user_id=filter_user_id
    )
    # Faturas sao vendas que carregam invoice_meta no campo notes.
    # Isso mantem separado o fluxo de venda comum/recibo termico.
    invoice_sales = [sale for sale in sales if controller._extract_invoice_meta(sale.notes)]

    if payment_status:
        invoice_sales = [s for s in invoice_sales if getattr(s, "payment_status", None) == payment_status]
    return invoice_sales


@router.get("/invoice-customers", response_model=List[schemas.PDVInvoiceCustomer])
def list_invoice_customers(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    terminal = controller.get_terminal_required(db, current_user.id)
    controller.require_terminal_permission(db, terminal.id, current_user.id, "can_sell")
    return controller.list_invoice_customers(db, terminal.id)


@router.post("/invoice-customers", response_model=schemas.PDVInvoiceCustomer)
def create_invoice_customer(
    payload: schemas.PDVInvoiceCustomerCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    terminal = controller.get_terminal_required(db, current_user.id)
    controller.require_terminal_permission(db, terminal.id, current_user.id, "can_sell")
    return controller.create_invoice_customer(db, payload, terminal.id, current_user.id)


@router.put("/invoice-customers/{customer_id}", response_model=schemas.PDVInvoiceCustomer)
def update_invoice_customer(
    customer_id: int,
    payload: schemas.PDVInvoiceCustomerUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    terminal = controller.get_terminal_required(db, current_user.id)
    controller.require_terminal_permission(db, terminal.id, current_user.id, "can_sell")
    return controller.update_invoice_customer(db, customer_id, payload, terminal.id)


@router.delete("/invoice-customers/{customer_id}")
def delete_invoice_customer(
    customer_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    terminal = controller.get_terminal_required(db, current_user.id)
    controller.require_terminal_permission(db, terminal.id, current_user.id, "can_sell")
    controller.delete_invoice_customer(db, customer_id, terminal.id)
    return {"message": "Invoice customer deleted"}


@router.post("/invoices/{invoice_id}/pay", response_model=schemas.PDVSale)
def pay_invoice(
    invoice_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    terminal = controller.get_terminal_required(db, current_user.id)
    controller.require_terminal_permission(db, terminal.id, current_user.id, "can_sell")
    sale = db.query(PDVSale).filter(
        PDVSale.id == invoice_id,
        PDVSale.terminal_id == terminal.id
    ).first()
    if not sale:
        raise HTTPException(status_code=404, detail="Invoice not found")
    if not controller._extract_invoice_meta(sale.notes):
        raise HTTPException(status_code=400, detail="This sale is not an invoice")
    return controller.mark_invoice_paid(db, invoice_id, terminal.id, current_user.id)


@router.post("/invoices/{invoice_id}/generate-receipt", response_model=schemas.PDVSale)
def generate_invoice_receipt(
    invoice_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    terminal = controller.get_terminal_required(db, current_user.id)
    controller.require_terminal_permission(db, terminal.id, current_user.id, "can_sell")
    return controller.mark_invoice_receipt_generated(db, invoice_id, terminal.id, current_user.id)

@router.get("/invoices/{invoice_id}/pdf")
def get_invoice_pdf(
    invoice_id: int,
    phone: Optional[str] = None,
    document_type: str = Query("invoice", pattern="^(invoice|receipt)$"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    terminal = controller.get_terminal_required(db, current_user.id)
    # Pode ver a própria venda ou, se admin, qualquer uma
    sale = db.query(PDVSale).filter(
        PDVSale.id == invoice_id,
        PDVSale.terminal_id == terminal.id
    ).first()
    if not sale:
        raise HTTPException(status_code=404, detail="Invoice not found")
    if not controller._extract_invoice_meta(sale.notes):
        raise HTTPException(status_code=400, detail="This sale is not an invoice")
    if not controller.is_terminal_admin(db, terminal.id, current_user.id) and sale.created_by != current_user.id:
        raise HTTPException(status_code=403, detail="Not allowed")

    items = db.query(PDVSaleItem).filter(PDVSaleItem.sale_id == sale.id).all()
    if document_type == "receipt":
        pdf_bytes = controller.generate_receipt_pdf(sale, terminal, items)
        filename = f"recibo-{sale.id}.pdf"
        whatsapp_caption = "Recibo SkyPDV"
    else:
        pdf_bytes = controller.generate_invoice_pdf(sale, terminal, items)
        filename = f"fatura-{sale.id}.pdf"
        whatsapp_caption = "Fatura SkyPDV"

    if phone:
        send_whatsapp_file(phone, filename, "application/pdf", pdf_bytes, caption=whatsapp_caption)

    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="{filename}"'}
    )

@router.get("/sales/{sale_id}", response_model=schemas.PDVSale)
def get_sale_details(
    sale_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Detalhes de uma venda específica
    - Caixas só podem ver suas próprias vendas
    - Admins podem ver todas as vendas do terminal
    """
    terminal = controller.get_terminal_required(db, current_user.id)
    if controller.is_terminal_admin(db, terminal.id, current_user.id):
        filter_user_id = None
    else:
        filter_user_id = current_user.id
    return controller.get_sale_details(db, sale_id, terminal.id, filter_user_id)

@router.post("/sales/{sale_id}/void", response_model=schemas.PDVSale)
def void_sale(
    sale_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Anular uma venda e estornar stock/caixa"""
    terminal = controller.get_terminal_required(db, current_user.id)
    return controller.void_sale(db, sale_id, terminal.id, current_user.id)

# ===================================================================
# Dashboard Endpoints
# ===================================================================

@router.get("/dashboard", response_model=schemas.DashboardStats)
def get_dashboard(
    user_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Estatísticas do dashboard PDV (Hoje/Mês/Top Produtos)
    - Caixas veem apenas suas próprias estatísticas
    - Admins veem estatísticas de todos os caixas
    - Se user_id for fornecido e usuário for admin, filtra por esse caixa específico
    """
    terminal = controller.get_terminal_required(db, current_user.id)
    if controller.is_terminal_admin(db, terminal.id, current_user.id):
        filter_user_id = user_id
    else:
        filter_user_id = current_user.id
    return controller.get_dashboard_stats(db, terminal.id, filter_user_id)

@router.get("/reports/sales-summary", response_model=schemas.SalesSummary)
def get_sales_report(
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    user_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Gerar relatório resumido de vendas para um periodo livre
    - Caixas veem apenas suas próprias vendas
    - Admins veem todas as vendas do terminal
    - Se user_id for fornecido e usuário for admin, filtra por esse caixa específico
    """
    terminal = controller.get_terminal_required(db, current_user.id)
    
    # Se não informar datas, assume mês atual
    if not start_date:
        start_date = datetime.utcnow().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    else:
        # Garantir que start_date comece no início do dia se não houver tempo
        if start_date.hour == 0 and start_date.minute == 0 and start_date.second == 0:
            start_date = start_date.replace(hour=0, minute=0, second=0, microsecond=0)

    if not end_date:
        end_date = datetime.utcnow()
    else:
        # Se end_date foi fornecido sem horas, assume fim do dia
        if end_date.hour == 0 and end_date.minute == 0 and end_date.second == 0:
            end_date = end_date.replace(hour=23, minute=59, second=59, microsecond=999999)
    
    # Se user_id foi fornecido e usuário é admin, usar esse user_id
    # Caso contrário, usar current_user.id (filtro automático)
    if controller.is_terminal_admin(db, terminal.id, current_user.id):
        filter_user_id = user_id
    else:
        filter_user_id = current_user.id
        
    return controller.get_sales_summary(db, terminal.id, start_date, end_date, filter_user_id)

@router.get("/reports/sales-summary.pdf")
def get_sales_report_pdf(
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    user_id: Optional[int] = None,
    product_scope: str = Query("all", pattern="^(all|beverages)$"),
    phone: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    terminal = controller.get_terminal_required(db, current_user.id)
 
    # Se não informar datas, assume mês atual
    if not start_date:
        start_date = datetime.utcnow().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    else:
        if start_date.hour == 0 and start_date.minute == 0 and start_date.second == 0:
            start_date = start_date.replace(hour=0, minute=0, second=0, microsecond=0)

    if not end_date:
        end_date = datetime.utcnow()
    else:
        if end_date.hour == 0 and end_date.minute == 0 and end_date.second == 0:
            end_date = end_date.replace(hour=23, minute=59, second=59, microsecond=999999)
 
    if controller.is_terminal_admin(db, terminal.id, current_user.id):
        filter_user_id = user_id
    else:
        filter_user_id = current_user.id
    summary = controller.get_sales_summary(db, terminal.id, start_date, end_date, filter_user_id)
 
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer
    currency = terminal.currency or "MT"
 
    def _fmt_dt(dt: Optional[datetime]) -> str:
        if not dt:
            return ""
        return dt.strftime("%d/%m/%Y %H:%M")
 
    def _fmt_date(dt: Optional[datetime]) -> str:
        if not dt:
            return ""
        return dt.strftime("%d/%m/%Y")

    def _fmt_2(v) -> str:
        if v is None:
            return f"0.00 {currency}"
        if isinstance(v, bool):
            return f"1.00 {currency}" if v else f"0.00 {currency}"
        try:
            val = float(v)
            return f"{val:,.2f} {currency}"
        except Exception:
            return f"{str(v)} {currency}"

    def _fmt_int(v) -> str:
        if v is None:
            return "0"
        try:
            return str(int(float(v)))
        except Exception:
            return str(v)

    beverage_filter = or_(
        func.lower(func.coalesce(PDVProduct.category, "")).like("%bebida%"),
        func.lower(func.coalesce(PDVProduct.category, "")).like("%drink%"),
        func.lower(func.coalesce(PDVProduct.name, "")).like("%sumo%"),
        func.lower(func.coalesce(PDVProduct.name, "")).like("%suco%"),
        func.lower(func.coalesce(PDVProduct.name, "")).like("%agua%"),
        func.lower(func.coalesce(PDVProduct.name, "")).like("%água%"),
        func.lower(func.coalesce(PDVProduct.name, "")).like("%refrigerante%"),
        func.lower(func.coalesce(PDVProduct.name, "")).like("%cerveja%"),
        func.lower(func.coalesce(PDVProduct.name, "")).like("%vinho%"),
        func.lower(func.coalesce(PDVProduct.name, "")).like("%whisky%"),
        func.lower(func.coalesce(PDVProduct.name, "")).like("%cafe%"),
        func.lower(func.coalesce(PDVProduct.name, "")).like("%café%"),
        func.lower(func.coalesce(PDVProduct.name, "")).like("%cha%"),
        func.lower(func.coalesce(PDVProduct.name, "")).like("%chá%"),
        func.lower(func.coalesce(PDVProduct.name, "")).like("%milkshake%"),
        func.lower(func.coalesce(PDVProduct.name, "")).like("%juice%"),
        func.lower(func.coalesce(PDVProduct.name, "")).like("%soda%"),
    )

    def _apply_product_scope(query):
        if product_scope == "beverages":
            return query.filter(beverage_filter)
        return query
 
    issued_at = datetime.utcnow()
    period_label = f"{_fmt_date(start_date)} até {_fmt_date(end_date)}"
 
    scope_label = "Apenas bebidas" if product_scope == "beverages" else "Todos os produtos"
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=36, rightMargin=36, topMargin=36, bottomMargin=36)
    styles = getSampleStyleSheet()
    story = []

    # Cabeçalho do Terminal
    story.append(Paragraph(terminal.name or "SkyPDV", styles["Title"]))
    if terminal.address:
        story.append(Paragraph(terminal.address, styles["Normal"]))
    story.append(Spacer(1, 12))

    story.append(Paragraph("Relatório: Vendas (Resumo)", styles["Heading2"]))
    story.append(Paragraph(f"Periodo: {period_label}", styles["Normal"]))
    story.append(Paragraph(f"Escopo dos produtos: {scope_label}", styles["Normal"]))
    story.append(Paragraph(f"Emitido em: {_fmt_dt(issued_at)} (Local)", styles["Normal"]))
    story.append(Spacer(1, 12))

    story.append(Paragraph("Resumo Financeiro", styles["Heading3"]))
    summary_table_data = [
        ["Métrica", "Valor"],
        ["Total de vendas", _fmt_int(summary.get("total_sales"))],
        ["Receita total", _fmt_2(summary.get("total_revenue"))],
        ["Custo total", _fmt_2(summary.get("total_cost"))],
        ["Lucro bruto", _fmt_2(summary.get("gross_profit"))],
        ["Ticket médio", _fmt_2(summary.get("average_sale_value"))],
        ["Itens vendidos", _fmt_int(summary.get("total_items_sold"))],
        ["Descontos", _fmt_2(summary.get("total_discounts"))],
        ["Impostos", _fmt_2(summary.get("total_taxes"))],
        [
            "Vendas anuladas",
            f"{_fmt_int(summary.get('voided_sales'))} (Total: {_fmt_2(summary.get('voided_amount'))})",
        ],
    ]
    summary_table = Table(summary_table_data, colWidths=[170, 340])
    summary_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F2F2F2")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ALIGN", (1, 1), (1, -1), "RIGHT"),
            ]
        )
    )
    story.append(summary_table)
    story.append(Spacer(1, 12))

    story.append(Paragraph("Pagamentos por Metodo", styles["Heading3"]))
    cash_total = summary.get("cash_sales") or 0
    card_total = summary.get("card_sales") or 0
    skywallet_total = summary.get("skywallet_sales") or 0
    mpesa_total = summary.get("mpesa_sales") or 0
    mixed_total = summary.get("mixed_sales") or 0
    payments_table_data = [
        ["Método", "Total"],
        ["Cash", _fmt_2(cash_total)],
        ["M-pesa", _fmt_2(mpesa_total)],
        ["E-Mola", _fmt_2(skywallet_total)],
        ["BCI POS", _fmt_2(card_total)],
        ["Misto", _fmt_2(mixed_total)],
        ["Total pagamentos", _fmt_2(cash_total + card_total + skywallet_total + mpesa_total + mixed_total)],
    ]
    def _has_value(v) -> bool:
        try:
            return float(v) != 0.0
        except Exception:
            return bool(v)

    payment_rows = [
        ("Cash", cash_total),
        ("M-pesa", mpesa_total),
        ("E-Mola", skywallet_total),
        ("BCI POS", card_total),
        ("Misto", mixed_total),
    ]
    visible_rows = [(name, total) for name, total in payment_rows if _has_value(total)]
    payments_table_data = [["Metodo", "Total"]]
    if not visible_rows:
        payments_table_data.append(["Sem pagamentos", _fmt_2(0)])
    else:
        for name, total in visible_rows:
            payments_table_data.append([name, _fmt_2(total)])
    payments_table_data.append(
        [
            "Total pagamentos",
            _fmt_2(cash_total + card_total + skywallet_total + mpesa_total + mixed_total),
        ]
    )

    payments_table = Table(payments_table_data, colWidths=[170, 340])
    payments_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F2F2F2")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("ALIGN", (1, 1), (1, -1), "RIGHT"),
            ]
        )
    )
    story.append(payments_table)
    story.append(Spacer(1, 12))

    story.append(Paragraph("Produtos mais vendidos (Top 30)", styles["Heading3"]))
    top_products = controller.get_top_products_report(db, terminal.id, start_date, end_date, limit=100, user_id=filter_user_id)
    if product_scope == "beverages":
        filtered_top_products = []
        for p in top_products:
            product_name = str(p.get("product_name") or "").lower()
            if any(
                keyword in product_name
                for keyword in [
                    "bebida",
                    "drink",
                    "sumo",
                    "suco",
                    "agua",
                    "água",
                    "refrigerante",
                    "cerveja",
                    "vinho",
                    "whisky",
                    "cafe",
                    "café",
                    "cha",
                    "chá",
                    "milkshake",
                    "juice",
                    "soda",
                ]
            ):
                filtered_top_products.append(p)
        top_products = filtered_top_products[:30]
    else:
        top_products = top_products[:30]
    items_table_data = [["Produto", "Qtd", "Receita", "Lucro"]]
    for p in top_products:
        items_table_data.append(
            [
                str(p.get("product_name") or ""),
                _fmt_int(p.get("quantity_sold") or 0),
                _fmt_2(p.get("revenue") or 0),
                _fmt_2(p.get("profit") or 0),
            ]
        )
    items_table = Table(items_table_data, colWidths=[240, 70, 100, 100])
    items_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F2F2F2")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
            ]
        )
    )
    story.append(items_table)
    story.append(Spacer(1, 12))

    # Lista completa de produtos vendidos no periodo com estoque atual
    story.append(Paragraph("Produtos vendidos no periodo", styles["Heading3"]))
    sold_products_query = (
        db.query(
            PDVProduct.id.label("product_id"),
            PDVProduct.name,
            func.sum(PDVSaleItem.quantity).label("qty"),
            func.max(PDVInventory.quantity).label("stock"),
        )
        .join(PDVSale, PDVSale.id == PDVSaleItem.sale_id)
        .join(PDVProduct, PDVProduct.id == PDVSaleItem.product_id)
        .outerjoin(
            PDVInventory,
            (PDVInventory.product_id == PDVProduct.id) & (PDVInventory.terminal_id == terminal.id),
        )
        .filter(PDVSale.terminal_id == terminal.id)
        .filter(PDVSale.created_at >= start_date)
        .filter(PDVSale.created_at <= end_date)
        .filter(PDVSale.status == "completed")
    )
    if filter_user_id:
        sold_products_query = sold_products_query.filter(PDVSale.created_by == filter_user_id)
    sold_products = (
        _apply_product_scope(sold_products_query)
        .group_by(PDVProduct.id, PDVProduct.name)
        .order_by(func.sum(PDVSaleItem.quantity).desc())
        .all()
    )
    product_movements = (
        _apply_product_scope(
            db.query(
                PDVProduct.id.label("product_id"),
                PDVStockMovement.movement_type,
                func.sum(PDVStockMovement.quantity).label("quantity"),
            )
            .join(PDVProduct, PDVProduct.id == PDVStockMovement.product_id)
            .filter(PDVStockMovement.terminal_id == terminal.id)
            .filter(PDVStockMovement.created_at >= start_date)
            .filter(PDVStockMovement.created_at <= end_date)
            .group_by(PDVProduct.id, PDVStockMovement.movement_type)
        )
        .all()
    )
    movement_by_product = {}
    for movement in product_movements:
        stats = movement_by_product.setdefault(movement.product_id, {"entries": 0, "exits": 0})
        amount = float(movement.quantity or 0)
        if movement.movement_type in (MovementType.IN, MovementType.RETURN):
            stats["entries"] += amount
        elif movement.movement_type in (MovementType.OUT, MovementType.SALE):
            stats["exits"] += abs(amount)
    sold_table_data = [["Produto", "Qtd vendida", "Entradas", "Saidas", "Estoque atual"]]
    for product in sold_products:
        movement_stats = movement_by_product.get(product.product_id, {"entries": 0, "exits": 0})
        sold_table_data.append(
            [
                str(product.name or ""),
                _fmt_int(product.qty or 0),
                _fmt_int(movement_stats["entries"]),
                _fmt_int(movement_stats["exits"]),
                _fmt_int(product.stock or 0),
            ]
        )
    if len(sold_table_data) == 1:
        sold_table_data.append(["Sem produtos vendidos", "0", "0", "0", "0"])
    sold_table = Table(sold_table_data, colWidths=[220, 75, 70, 70, 75])
    sold_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F2F2F2")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
            ]
        )
    )
    story.append(sold_table)
    story.append(Spacer(1, 12))

    story.append(Paragraph("Movimento de Produtos (Entradas e Saídas)", styles["Heading3"]))
    movements = (
        db.query(PDVStockMovement.movement_type, func.sum(PDVStockMovement.quantity))
        .filter(PDVStockMovement.terminal_id == terminal.id)
        .filter(PDVStockMovement.created_at >= start_date)
        .filter(PDVStockMovement.created_at <= end_date)
        .group_by(PDVStockMovement.movement_type)
        .all()
    )
    move_sums = {mt: (qty or 0) for mt, qty in movements}
    entries = (move_sums.get(MovementType.IN, 0) or 0) + (move_sums.get(MovementType.RETURN, 0) or 0)
    exits = abs(move_sums.get(MovementType.OUT, 0) or 0) + abs(move_sums.get(MovementType.SALE, 0) or 0)
    transfers = move_sums.get(MovementType.TRANSFER, 0) or 0
    adjustments = move_sums.get(MovementType.ADJUSTMENT, 0) or 0
    total_movements = entries + exits + abs(transfers) + abs(adjustments)
    stock_table_data = [
        ["Tipo", "Quantidade"],
        ["Entradas (IN/RETURN)", _fmt_int(entries)],
        ["Saídas (OUT/SALE)", _fmt_int(exits)],
        ["Transferências", _fmt_int(transfers)],
        ["Ajustes", _fmt_int(adjustments)],
        ["Total movimentado", _fmt_int(total_movements)],
    ]
    stock_table = Table(stock_table_data, colWidths=[240, 270])
    stock_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F2F2F2")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("ALIGN", (1, 1), (1, -1), "RIGHT"),
            ]
        )
    )
    story.append(stock_table)
    story.append(Spacer(1, 12))

    story.append(Paragraph("Movimentos por Produto", styles["Heading3"]))
    product_movements = (
        db.query(
            PDVProduct.name,
            PDVStockMovement.movement_type,
            func.sum(PDVStockMovement.quantity),
        )
        .join(PDVProduct, PDVProduct.id == PDVStockMovement.product_id)
        .filter(PDVStockMovement.terminal_id == terminal.id)
        .filter(PDVStockMovement.created_at >= start_date)
        .filter(PDVStockMovement.created_at <= end_date)
        .group_by(PDVProduct.name, PDVStockMovement.movement_type)
        .all()
    )
    per_product = {}
    for name, movement_type, qty in product_movements:
        entry = per_product.setdefault(
            name or "Sem nome",
            {"entries": 0, "exits": 0, "transfers": 0, "adjustments": 0},
        )
        amount = float(qty or 0)
        if movement_type in (MovementType.IN, MovementType.RETURN):
            entry["entries"] += amount
        elif movement_type in (MovementType.OUT, MovementType.SALE):
            entry["exits"] += abs(amount)
        elif movement_type == MovementType.TRANSFER:
            entry["transfers"] += abs(amount)
        elif movement_type == MovementType.ADJUSTMENT:
            entry["adjustments"] += abs(amount)

    if per_product:
        rows = [["Produto", "Entradas", "Saídas", "Total mov."]]
        items = []
        for name, data in per_product.items():
            total = data["entries"] + data["exits"] + data["transfers"] + data["adjustments"]
            items.append((name, data["entries"], data["exits"], total))
        items.sort(key=lambda x: x[3], reverse=True)
        for name, entries_val, exits_val, total_val in items[:50]:
            rows.append(
                [
                    str(name),
                    _fmt_int(entries_val),
                    _fmt_int(exits_val),
                    _fmt_int(total_val),
                ]
            )
        product_table = Table(rows, colWidths=[240, 70, 70, 70])
        product_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F2F2F2")),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
                ]
            )
        )
        story.append(product_table)
    else:
        story.append(Paragraph("Nenhum movimento de produto no periodo.", styles["Normal"]))

    doc.build(story)
    pdf_bytes = buffer.getvalue()

    filename = f"Relatorio_Vendas_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.pdf"
    headers = {"Content-Disposition": f"attachment; filename=\"{filename}\""}

    if phone:
        caption = f"Relatório de vendas {start_date.strftime('%d/%m/%Y')} a {end_date.strftime('%d/%m/%Y')} (SkyPDV)."
        send_whatsapp_file(phone, filename, "application/pdf", pdf_bytes, caption=caption)
        send_whatsapp_text(phone, caption)

    return StreamingResponse(io.BytesIO(pdf_bytes), media_type="application/pdf", headers=headers)


@router.get("/reports/sales-summary.xlsx")
def get_sales_report_excel(
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    user_id: Optional[int] = None,
    phone: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Exporta o resumo de vendas em Excel (XLSX).
    Usa os mesmos filtros do relatório PDF.
    """
    terminal = controller.get_terminal_required(db, current_user.id)

    # Datas padrão: mês atual
    if not start_date:
        start_date = datetime.utcnow().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    else:
        if start_date.hour == 0 and start_date.minute == 0 and start_date.second == 0:
            start_date = start_date.replace(hour=0, minute=0, second=0, microsecond=0)

    if not end_date:
        end_date = datetime.utcnow()
    else:
        if end_date.hour == 0 and end_date.minute == 0 and end_date.second == 0:
            end_date = end_date.replace(hour=23, minute=59, second=59, microsecond=999999)

    filter_user_id = user_id if controller.is_terminal_admin(db, terminal.id, current_user.id) else current_user.id
    summary = controller.get_sales_summary(db, terminal.id, start_date, end_date, filter_user_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumo de Vendas"

    cash_total = summary.get("cash_sales") or 0
    card_total = summary.get("card_sales") or 0
    skywallet_total = summary.get("skywallet_sales") or 0
    mpesa_total = summary.get("mpesa_sales") or 0
    mixed_total = summary.get("mixed_sales") or 0
    total_payments = cash_total + card_total + skywallet_total + mpesa_total + mixed_total

    rows = [
        ("Periodo Inicio", start_date.strftime("%d/%m/%Y %H:%M")),
        ("Periodo Fim", end_date.strftime("%d/%m/%Y %H:%M")),
        ("Total Vendas", summary["total_sales"]),
        ("Receita Bruta", summary["total_revenue"]),
        ("Custo Total", summary["total_cost"]),
        ("Lucro Bruto", summary["gross_profit"]),
        ("Ticket Médio", summary["average_sale_value"]),
        ("Itens Vendidos", summary["total_items_sold"]),
        ("Descontos", summary["total_discounts"]),
        ("Impostos", summary["total_taxes"]),
        ("Vendas Cash", cash_total),
        ("Vendas M-pesa", mpesa_total),
        ("Vendas E-Mola", skywallet_total),
        ("Vendas BCI POS", card_total),
        ("Vendas Misto", mixed_total),
        ("Total pagamentos", total_payments),
        ("Vendas Anuladas", summary["voided_sales"]),
        ("Valor Anulado", summary["voided_amount"]),
    ]

    ws.append(["Métrica", "Valor"])
    def _has_value(v) -> bool:
        try:
            return float(v) != 0.0
        except Exception:
            return bool(v)

    payment_row_names = {
        "Vendas Cash",
        "Vendas M-pesa",
        "Vendas E-Mola",
        "Vendas BCI POS",
        "Vendas Misto",
    }
    payment_rows = [
        ("Vendas Cash", cash_total),
        ("Vendas M-pesa", mpesa_total),
        ("Vendas E-Mola", skywallet_total),
        ("Vendas BCI POS", card_total),
        ("Vendas Misto", mixed_total),
    ]
    visible_payments = [(name, total) for name, total in payment_rows if _has_value(total)]
    cleaned_rows = []
    for name, value in rows:
        if name in payment_row_names:
            continue
        if name == "Total pagamentos":
            if not visible_payments:
                cleaned_rows.append(("Sem pagamentos", 0))
            else:
                cleaned_rows.extend(visible_payments)
            cleaned_rows.append((name, value))
            continue
        cleaned_rows.append((name, value))
    rows = cleaned_rows

    for name, value in rows:
        ws.append([name, value])

    ws.append([])
    ws.append(["Movimento de Produtos (Entradas e Saídas)", "Quantidade"])
    movements = (
        db.query(PDVStockMovement.movement_type, func.sum(PDVStockMovement.quantity))
        .filter(PDVStockMovement.terminal_id == terminal.id)
        .filter(PDVStockMovement.created_at >= start_date)
        .filter(PDVStockMovement.created_at <= end_date)
        .group_by(PDVStockMovement.movement_type)
        .all()
    )
    move_sums = {mt: (qty or 0) for mt, qty in movements}
    entries = (move_sums.get(MovementType.IN, 0) or 0) + (move_sums.get(MovementType.RETURN, 0) or 0)
    exits = abs(move_sums.get(MovementType.OUT, 0) or 0) + abs(move_sums.get(MovementType.SALE, 0) or 0)
    transfers = move_sums.get(MovementType.TRANSFER, 0) or 0
    adjustments = move_sums.get(MovementType.ADJUSTMENT, 0) or 0
    total_movements = entries + exits + abs(transfers) + abs(adjustments)
    ws.append(["Entradas (IN/RETURN)", entries])
    ws.append(["Saídas (OUT/SALE)", exits])
    ws.append(["Transferências", transfers])
    ws.append(["Ajustes", adjustments])
    ws.append(["Total movimentado", total_movements])

    from openpyxl.styles import PatternFill, Font

    ws_products = wb.create_sheet("Movimentos Produtos")
    ws_products.append(["Produto", "Entradas", "Saidas", "Estoque atual"])
    product_movements = (
        db.query(
            PDVProduct.name,
            PDVStockMovement.movement_type,
            func.sum(PDVStockMovement.quantity),
        )
        .join(PDVProduct, PDVProduct.id == PDVStockMovement.product_id)
        .filter(PDVStockMovement.terminal_id == terminal.id)
        .filter(PDVStockMovement.created_at >= start_date)
        .filter(PDVStockMovement.created_at <= end_date)
        .group_by(PDVProduct.name, PDVStockMovement.movement_type)
        .all()
    )
    per_product = {}
    for name, movement_type, qty in product_movements:
        entry = per_product.setdefault(
            name or "Sem nome",
            {"entries": 0, "exits": 0, "transfers": 0, "adjustments": 0},
        )
        amount = float(qty or 0)
        if movement_type in (MovementType.IN, MovementType.RETURN):
            entry["entries"] += amount
        elif movement_type in (MovementType.OUT, MovementType.SALE):
            entry["exits"] += abs(amount)
        elif movement_type == MovementType.TRANSFER:
            entry["transfers"] += abs(amount)
        elif movement_type == MovementType.ADJUSTMENT:
            entry["adjustments"] += abs(amount)
    product_current_stock = {
        row.product_id: float(row.current_stock or 0)
        for row in (
            db.query(
                PDVInventory.product_id,
                func.sum(PDVInventory.quantity).label("current_stock"),
            )
            .filter(PDVInventory.terminal_id == terminal.id)
            .group_by(PDVInventory.product_id)
            .all()
        )
    }
    items = []
    for name, data in per_product.items():
        total = data["entries"] + data["exits"] + data["transfers"] + data["adjustments"]
        items.append((name, data["entries"], data["exits"], total))
    # Produtos vendidos (saídas > 0) aparecem primeiro.
    items.sort(
        key=lambda x: (
            0 if x[2] > 0 else 1,
            -x[2],
            str(x[0]).lower(),
        )
    )

    sold_fill = PatternFill(fill_type="solid", fgColor="E8F5E9")
    sold_font = Font(color="1B5E20")

    for name, entries_val, exits_val, _total_val in items:
        current_stock = 0.0
        for row in (
            db.query(PDVProduct.id)
            .filter(PDVProduct.terminal_id == terminal.id)
            .filter(PDVProduct.name == name)
            .all()
        ):
            current_stock += product_current_stock.get(row.id, 0.0)
        ws_products.append([name, entries_val, exits_val, current_stock])
        if float(exits_val or 0) > 0:
            current_row = ws_products.max_row
            for col in range(1, 5):
                cell = ws_products.cell(row=current_row, column=col)
                cell.fill = sold_fill
                cell.font = sold_font

    # Auto ajuste de largura
    for col in range(1, 3):
        max_length = 0
        col_letter = get_column_letter(col)
        for cell in ws[col_letter]:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"sales-summary-{start_date.strftime('%Y%m%d')}-{end_date.strftime('%Y%m%d')}.xlsx"
    headers = {
        "Content-Disposition": f'attachment; filename="{filename}"',
        "Access-Control-Expose-Headers": "Content-Disposition",
    }

    if phone:
        caption = f"Relatório de vendas (Excel) {start_date.strftime('%d/%m/%Y')} a {end_date.strftime('%d/%m/%Y')} (SkyPDV)."
        send_whatsapp_file(phone, filename, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", buffer.getvalue(), caption=caption)
        send_whatsapp_text(phone, caption)

    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)
@router.get("/reports/products.pdf")
def get_products_report_pdf(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    terminal = controller.get_terminal_required(db, current_user.id)

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from decimal import Decimal
    from models import PDVProduct

    def _fmt_dt(dt: Optional[datetime]) -> str:
        if not dt:
            return ""
        return dt.strftime("%d/%m/%Y %H:%M")

    def _fmt_date(dt: Optional[datetime]) -> str:
        if not dt:
            return ""
        return dt.strftime("%d/%m/%Y")

    def _fmt_2(v) -> str:
        if v is None:
            return "0.00"
        if isinstance(v, bool):
            return "1.00" if v else "0.00"
        if isinstance(v, int):
            return f"{v:.2f}"
        if isinstance(v, Decimal):
            return f"{v:.2f}"
        try:
            return f"{float(v):.2f}"
        except Exception:
            return str(v)

    issued_at = datetime.utcnow()

    products = (
        db.query(PDVProduct)
        .filter(PDVProduct.terminal_id == terminal.id)
        .filter(PDVProduct.is_active == True)
        .order_by(PDVProduct.name.asc())
        .all()
    )

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=36, rightMargin=36, topMargin=36, bottomMargin=36)
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph("Relatório: Produtos e Estoque", styles["Title"]))
    story.append(Paragraph(f"Emitido em: {_fmt_dt(issued_at)} (UTC)", styles["Normal"]))
    story.append(Spacer(1, 12))

    table_data = [["Produto", "SKU", "Estoque", "Preço"]]
    for p in products:
        inv_qty = None
        if getattr(p, "track_stock", False):
            inv = getattr(p, "inventory", None)
            inv_qty = getattr(inv, "quantity", None) if inv else None

        table_data.append(
            [
                str(getattr(p, "name", "") or ""),
                str(getattr(p, "sku", "") or ""),
                _fmt_2(inv_qty) if getattr(p, "track_stock", False) else "-",
                _fmt_2(getattr(p, "price", None)),
            ]
        )

    products_table = Table(table_data, colWidths=[260, 90, 70, 90], repeatRows=1)
    products_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F2F2F2")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ALIGN", (2, 1), (3, -1), "RIGHT"),
            ]
        )
    )

    story.append(products_table)

    doc.build(story)
    pdf_bytes = buffer.getvalue()

    filename = f"products_{_fmt_date(issued_at)}.pdf"
    headers = {"Content-Disposition": f"attachment; filename=\"{filename}\""}
    return StreamingResponse(buffer, media_type="application/pdf", headers=headers)

@router.get("/reports/periodic", response_model=schemas.SalesSummary)
def get_periodic_sales_report(
    period: str = Query(..., description="Tipo de periodo: day, month, year"),
    date: str = Query(..., description="Data no formato AAAA-MM-DD, AAAA-MM ou AAAA"),
    user_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Relatório simplificado por Dia, Mês ou Ano.
    - Ex: period='day', date='2024-01-21'
    - Ex: period='month', date='2024-01'
    - Ex: period='year', date='2024'
    - Caixas veem apenas suas próprias vendas
    - Admins veem todas as vendas do terminal
    - Se user_id for fornecido e usuário for admin, filtra por esse caixa específico
    """
    terminal = controller.get_terminal_required(db, current_user.id)
    # Se user_id foi fornecido e usuário é admin, usar esse user_id
    filter_user_id = user_id if (user_id and controller.is_terminal_admin(db, terminal.id, current_user.id)) else current_user.id
    return controller.get_periodic_report(db, terminal.id, period, date, filter_user_id)

@router.get("/reports/detailed-monthly", response_model=schemas.DetailedMonthlyReport)
def get_detailed_monthly_report(
    year: int = Query(..., description="Ano (ex: 2024)"),
    month: int = Query(..., description="Mês (1-12)"),
    user_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Relatório mensal detalhado com breakdown diário, top produtos, categorias, etc.
    - Caixas veem apenas suas próprias vendas
    - Admins veem todas as vendas do terminal
    - Se user_id for fornecido e usuário for admin, filtra por esse caixa específico
    """
    terminal = controller.get_terminal_required(db, current_user.id)
    filter_user_id = user_id if (user_id and controller.is_terminal_admin(db, terminal.id, current_user.id)) else current_user.id
    return controller.get_detailed_monthly_report(db, terminal.id, year, month, filter_user_id)

@router.get("/reports/detailed-yearly", response_model=schemas.DetailedYearlyReport)
def get_detailed_yearly_report(
    year: int = Query(..., description="Ano (ex: 2024)"),
    user_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Relatório anual detalhado com breakdown mensal, comparação, tendências, etc.
    - Caixas veem apenas suas próprias vendas
    - Admins veem todas as vendas do terminal
    - Se user_id for fornecido e usuário for admin, filtra por esse caixa específico
    """
    terminal = controller.get_terminal_required(db, current_user.id)
    filter_user_id = user_id if (user_id and controller.is_terminal_admin(db, terminal.id, current_user.id)) else current_user.id
    return controller.get_detailed_yearly_report(db, terminal.id, year, filter_user_id)

@router.get("/reports/top-products", response_model=List[schemas.TopProduct])
def get_top_products_report(
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    limit: int = Query(20, description="Número de produtos a retornar"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Relatorio de produtos mais vendidos em um periodo.
    Se não informar datas, assume mês atual.
    """
    terminal = controller.get_terminal_required(db, current_user.id)
    if not start_date:
        start_date = datetime.utcnow().replace(day=1, hour=0, minute=0, second=0)
    if not end_date:
        end_date = datetime.utcnow()
    return controller.get_top_products_report(db, terminal.id, start_date, end_date, limit)

@router.get("/reports/sales-by-day", response_model=List[schemas.SalesByPeriod])
def get_sales_by_day(
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    user_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Breakdown de vendas por dia em um periodo.
    Útil para gráficos de tendência diária.
    Se não informar datas, assume mês atual.
    """
    terminal = controller.get_terminal_required(db, current_user.id)
    if not start_date:
        start_date = datetime.utcnow().replace(day=1, hour=0, minute=0, second=0)
    if not end_date:
        end_date = datetime.utcnow()
    if controller.is_terminal_admin(db, terminal.id, current_user.id):
        filter_user_id = user_id
    else:
        filter_user_id = current_user.id
    return controller.get_sales_by_day(db, terminal.id, start_date, end_date, filter_user_id)

# ===================================================================
# Categories Endpoints
# ===================================================================

@router.get("/categories-list", response_model=List[schemas.PDVCategory])
def list_categories_full(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Listar todas as categorias cadastradas"""
    terminal = controller.get_terminal_required(db, current_user.id)
    return controller.get_categories_list(db, terminal.id)

@router.post("/categories-list", response_model=schemas.PDVCategory)
def create_category(
    category: schemas.PDVCategoryCreate,
    is_global: bool = False,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Criar nova categoria (pessoal ou global)"""
    terminal = controller.get_terminal_required(db, current_user.id)
    return controller.create_category(db, category, terminal.id, current_user.id, is_global)

@router.post("/categories-list/{category_id}/adopt", response_model=schemas.PDVCategory)
def adopt_category(
    category_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Adotar uma categoria global para o seu terminal"""
    terminal = controller.get_terminal_required(db, current_user.id)
    return controller.adopt_category(db, category_id, terminal.id, current_user.id)

@router.put("/categories-list/{category_id}", response_model=schemas.PDVCategory)
def update_category(
    category_id: int,
    updates: schemas.PDVCategoryUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Atualizar categoria"""
    terminal = controller.get_terminal_required(db, current_user.id)
    return controller.update_category(db, category_id, updates, terminal.id)

@router.delete("/categories-list/{category_id}")
def delete_category(
    category_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Desativar categoria"""
    terminal = controller.get_terminal_required(db, current_user.id)
    return controller.delete_category(db, category_id, terminal.id)

# ===================================================================
# Payment Methods Endpoints
# ===================================================================

@router.get("/payment-methods", response_model=List[schemas.PDVPaymentMethod])
def list_payment_methods(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Listar todos os métodos de pagamento cadastrados"""
    terminal = controller.get_terminal_required(db, current_user.id)
    return controller.get_payment_methods_list(db, terminal.id)

@router.post("/payment-methods", response_model=schemas.PDVPaymentMethod)
def create_payment_method(
    method: schemas.PDVPaymentMethodCreate,
    is_global: bool = False,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Criar novo método de pagamento (pessoal ou global)"""
    terminal = controller.get_terminal_required(db, current_user.id)
    return controller.create_payment_method(db, method, terminal.id, current_user.id, is_global)

@router.post("/payment-methods/{method_id}/adopt", response_model=schemas.PDVPaymentMethod)
def adopt_payment_method(
    method_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Adotar um método de pagamento global para o seu terminal"""
    terminal = controller.get_terminal_required(db, current_user.id)
    return controller.adopt_payment_method(db, method_id, terminal.id, current_user.id)

@router.put("/payment-methods/{method_id}", response_model=schemas.PDVPaymentMethod)
def update_payment_method(
    method_id: int,
    updates: schemas.PDVPaymentMethodUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Atualizar método de pagamento"""
    terminal = controller.get_terminal_required(db, current_user.id)
    return controller.update_payment_method(db, method_id, updates, terminal.id)

@router.delete("/payment-methods/{method_id}")
def delete_payment_method(
    method_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Desativar método de pagamento"""
    terminal = controller.get_terminal_required(db, current_user.id)
    return controller.delete_payment_method(db, method_id, terminal.id)

# ===================================================================
# Finance Endpoints (apenas admin do terminal: dono ou role ADMIN)
# ===================================================================

def _require_terminal_finance_admin(db: Session, terminal, current_user_id: int) -> None:
    if not controller.is_terminal_admin(db, terminal.id, current_user_id):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Apenas administradores do terminal podem aceder às finanças.",
        )


@router.get("/finance/expense-categories", response_model=List[schemas.PDVExpenseCategory])
def list_expense_categories(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    terminal = controller.get_terminal_required(db, current_user.id)
    _require_terminal_finance_admin(db, terminal, current_user.id)
    return controller.get_expense_categories_list(db, terminal.id)


@router.post("/finance/expense-categories", response_model=schemas.PDVExpenseCategory)
def create_expense_category(
    category: schemas.PDVExpenseCategoryCreate,
    is_global: bool = False,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    terminal = controller.get_terminal_required(db, current_user.id)
    _require_terminal_finance_admin(db, terminal, current_user.id)
    return controller.create_expense_category(db, category, terminal.id, current_user.id, is_global)


@router.put("/finance/expense-categories/{category_id}", response_model=schemas.PDVExpenseCategory)
def update_expense_category(
    category_id: int,
    updates: schemas.PDVExpenseCategoryUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    terminal = controller.get_terminal_required(db, current_user.id)
    _require_terminal_finance_admin(db, terminal, current_user.id)
    return controller.update_expense_category(db, category_id, updates, terminal.id)


@router.delete("/finance/expense-categories/{category_id}")
def delete_expense_category(
    category_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    terminal = controller.get_terminal_required(db, current_user.id)
    _require_terminal_finance_admin(db, terminal, current_user.id)
    return controller.delete_expense_category(db, category_id, terminal.id)


@router.get("/finance/expenses", response_model=List[schemas.PDVExpense])
def list_expenses(
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    category_id: Optional[int] = None,
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    terminal = controller.get_terminal_required(db, current_user.id)
    _require_terminal_finance_admin(db, terminal, current_user.id)
    return controller.get_expenses(db, terminal.id, start_date, end_date, category_id, skip, limit)


@router.post("/finance/expenses", response_model=schemas.PDVExpense)
def create_expense(
    expense: schemas.PDVExpenseCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    terminal = controller.get_terminal_required(db, current_user.id)
    _require_terminal_finance_admin(db, terminal, current_user.id)
    return controller.create_expense(db, expense, terminal.id, current_user.id)


@router.put("/finance/expenses/{expense_id}", response_model=schemas.PDVExpense)
def update_expense(
    expense_id: int,
    updates: schemas.PDVExpenseUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    terminal = controller.get_terminal_required(db, current_user.id)
    _require_terminal_finance_admin(db, terminal, current_user.id)
    return controller.update_expense(db, expense_id, updates, terminal.id)


@router.delete("/finance/expenses/{expense_id}")
def delete_expense(
    expense_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    terminal = controller.get_terminal_required(db, current_user.id)
    _require_terminal_finance_admin(db, terminal, current_user.id)
    return controller.delete_expense(db, expense_id, terminal.id)


@router.get("/finance/summary", response_model=schemas.FinancialSummary)
def get_finance_summary(
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    user_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    terminal = controller.get_terminal_required(db, current_user.id)
    _require_terminal_finance_admin(db, terminal, current_user.id)
    if not start_date:
        start_date = datetime.utcnow().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    if not end_date:
        end_date = datetime.utcnow()
    filter_user_id = user_id
    return controller.get_financial_summary(db, terminal.id, start_date, end_date, filter_user_id)


@router.get("/finance/tax-summary", response_model=schemas.PDVTaxSummary)
def get_finance_tax_summary(
    year: int,
    month: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    terminal = controller.get_terminal_required(db, current_user.id)
    _require_terminal_finance_admin(db, terminal, current_user.id)
    return controller.get_tax_summary(db, terminal.id, year, month)


@router.put("/finance/tax-summary", response_model=schemas.PDVTaxSummary)
def update_finance_tax_summary(
    year: int,
    month: int,
    payload: schemas.PDVTaxSummaryUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    terminal = controller.get_terminal_required(db, current_user.id)
    _require_terminal_finance_admin(db, terminal, current_user.id)
    return controller.update_tax_summary(db, terminal.id, year, month, current_user.id, payload)

@router.get("/finance/summary.pdf")
def get_finance_summary_pdf(
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    user_id: Optional[int] = None,
    phone: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    terminal = controller.get_terminal_required(db, current_user.id)
    _require_terminal_finance_admin(db, terminal, current_user.id)
    if not start_date:
        start_date = datetime.utcnow().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    if not end_date:
        end_date = datetime.utcnow()
    filter_user_id = user_id

    summary = controller.get_financial_summary(db, terminal.id, start_date, end_date, filter_user_id)

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=36, rightMargin=36, topMargin=36, bottomMargin=36)
    styles = getSampleStyleSheet()
    currency = terminal.currency or "MT"

    def fmt_amt(v):
        try:
            return f"{float(v):,.2f} {currency}"
        except Exception:
            return f"{v} {currency}"

    story = []
    title = f"Resumo Financeiro - {terminal.name or 'SkyPDV'}"
    story.append(Paragraph(title, styles["Title"]))
    story.append(Spacer(1, 8))
    story.append(
        Paragraph(
            f"Periodo: {start_date.strftime('%d/%m/%Y')} a {end_date.strftime('%d/%m/%Y')}",
            styles["Normal"],
        )
    )
    story.append(Spacer(1, 12))

    # Totais
    totals_data = [
        ["Entradas (vendas)", fmt_amt(summary["gross_revenue"])],
        ["Lucro bruto", fmt_amt(summary["gross_profit"])],
        ["Saídas (despesas)", fmt_amt(summary["total_expenses"])],
        ["Lucro líquido", fmt_amt(summary["net_profit"])],
        ["Nº vendas", str(summary["sales_count"])],
        ["Nº despesas", str(summary["expenses_count"])],
    ]
    totals_table = Table([["Métrica", "Valor"]] + totals_data, colWidths=[220, 200])
    totals_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F2F2F2")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("ALIGN", (1, 1), (1, -1), "RIGHT"),
            ]
        )
    )
    story.append(totals_table)
    story.append(Spacer(1, 12))

    # Breakdown
    breakdown = summary.get("expense_breakdown") or []
    story.append(Paragraph("Despesas por categoria", styles["Heading3"]))
    if breakdown:
        rows = [["Categoria", "Valor"]]
        for item in breakdown:
            rows.append(
                [
                    item.get("category_name") or "Sem categoria",
                    fmt_amt(item.get("total_amount") or 0),
                ]
            )
        table = Table(rows, colWidths=[260, 160])
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F2F2F2")),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                    ("ALIGN", (1, 1), (1, -1), "RIGHT"),
                ]
            )
        )
        story.append(table)
    else:
        story.append(Paragraph("Sem despesas no periodo.", styles["Normal"]))

    doc.build(story)
    buffer.seek(0)
    filename = f"financeiro_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.pdf"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}

    if phone:
        caption = f"Resumo financeiro {start_date.strftime('%d/%m/%Y')} a {end_date.strftime('%d/%m/%Y')} (SkyPDV)."
        send_whatsapp_file(phone, filename, "application/pdf", pdf_bytes, caption=caption)
        send_whatsapp_text(phone, caption)

    return StreamingResponse(io.BytesIO(pdf_bytes), media_type="application/pdf", headers=headers)


@router.get("/reports/stock-day.pdf")
def get_stock_day_report_pdf(
    date: Optional[datetime] = None,
    product_scope: str = Query("all", pattern="^(all|beverages|important)$"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    terminal = controller.get_terminal_required(db, current_user.id)

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from decimal import Decimal
    from models import PDVProduct

    def _fmt_dt(dt: Optional[datetime]) -> str:
        if not dt:
            return ""
        return dt.strftime("%d/%m/%Y %H:%M")

    def _fmt_num(v, digits: int = 3) -> str:
        if v is None:
            return f"{0:.{digits}f}"
        if isinstance(v, bool):
            return f"{1 if v else 0:.{digits}f}"
        if isinstance(v, int):
            return f"{v:.{digits}f}"
        if isinstance(v, Decimal):
            return f"{v:.{digits}f}"
        try:
            return f"{float(v):.{digits}f}"
        except Exception:
            return str(v)

    def _fmt_qty(v) -> str:
        """Format quantity without forcing trailing .000 for integers."""
        try:
            value = float(v or 0)
        except Exception:
            return str(v)
        if value.is_integer():
            return str(int(value))
        text = f"{value:.3f}".rstrip("0").rstrip(".")
        return text or "0"

    report_day = date or datetime.utcnow()
    start_date = report_day.replace(hour=0, minute=0, second=0, microsecond=0)
    end_date = report_day.replace(hour=23, minute=59, second=59, microsecond=999999)
    issued_at = datetime.utcnow()
    if product_scope == "beverages":
        scope_label = "Apenas bebidas"
    elif product_scope == "important":
        scope_label = "Apenas produtos com estoque controlado"
    else:
        scope_label = "Todos os produtos"

    def _is_beverage_product(product: PDVProduct) -> bool:
        category = str(getattr(product, "category", "") or "").lower()
        name = str(getattr(product, "name", "") or "").lower()
        keywords = [
            "bebida", "drink", "sumo", "suco", "agua", "água", "refrigerante",
            "cerveja", "vinho", "whisky", "cafe", "café", "cha", "chá",
            "milkshake", "juice", "soda",
        ]
        return any(k in category for k in keywords) or any(k in name for k in keywords)

    inventory_rows = (
        db.query(PDVInventory, PDVProduct)
        .join(PDVProduct, PDVProduct.id == PDVInventory.product_id)
        .filter(PDVInventory.terminal_id == terminal.id)
        .filter(PDVProduct.is_active == True)
        .filter(PDVProduct.track_stock == True)
        .order_by(PDVProduct.name.asc(), PDVInventory.storage_location.asc())
        .all()
    )

    movement_rows = (
        db.query(PDVStockMovement, PDVProduct)
        .join(PDVProduct, PDVProduct.id == PDVStockMovement.product_id)
        .filter(PDVStockMovement.terminal_id == terminal.id)
        .filter(PDVStockMovement.created_at >= start_date)
        .filter(PDVStockMovement.created_at <= end_date)
        .order_by(PDVStockMovement.created_at.desc())
        .all()
    )

    if product_scope == "beverages":
        inventory_rows = [(inv, prod) for inv, prod in inventory_rows if _is_beverage_product(prod)]
        movement_rows = [(mov, prod) for mov, prod in movement_rows if _is_beverage_product(prod)]
    elif product_scope == "important":
        movement_rows = [(mov, prod) for mov, prod in movement_rows if bool(getattr(prod, "track_stock", False))]

    totals = {"entries": 0.0, "exits": 0.0, "adjustments": 0.0, "transfers": 0.0, "sales": 0.0}
    for movement, _product in movement_rows:
        qty = abs(float(movement.quantity or 0))
        if movement.movement_type in (MovementType.IN, MovementType.RETURN):
            totals["entries"] += qty
        elif movement.movement_type == MovementType.SALE:
            totals["sales"] += qty
            totals["exits"] += qty
        elif movement.movement_type == MovementType.OUT:
            totals["exits"] += qty
        elif movement.movement_type == MovementType.ADJUSTMENT:
            totals["adjustments"] += qty
        elif movement.movement_type == MovementType.TRANSFER:
            totals["transfers"] += qty

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=28, rightMargin=28, topMargin=28, bottomMargin=28)
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph("Relatorio: Stock do Dia", styles["Title"]))
    story.append(Paragraph(terminal.name or "SkyPDV", styles["Heading2"]))
    story.append(Paragraph(f"Data do stock: {start_date.strftime('%d/%m/%Y')}", styles["Normal"]))
    story.append(Paragraph(f"Escopo dos produtos: {scope_label}", styles["Normal"]))
    story.append(Paragraph(f"Emitido em: {_fmt_dt(issued_at)}", styles["Normal"]))
    story.append(Spacer(1, 12))

    summary_rows = [
        ["Resumo", "Quantidade"],
        ["Entradas", _fmt_num(totals["entries"], 0)],
        ["Saidas", _fmt_num(totals["exits"], 0)],
        ["Vendido", _fmt_num(totals["sales"], 0)],
        ["Ajustes", _fmt_num(totals["adjustments"], 0)],
        ["Transferencias", _fmt_num(totals["transfers"], 0)],
        ["Movimentos do dia", str(len(movement_rows))],
        ["Itens em estoque", str(len(inventory_rows))],
    ]
    summary_table = Table(summary_rows, colWidths=[260, 220])
    summary_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F2F2F2")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("ALIGN", (1, 1), (1, -1), "RIGHT"),
            ]
        )
    )
    story.append(summary_table)
    story.append(Spacer(1, 12))

    product_summary = {}
    for inventory, product in inventory_rows:
        product_name = str(product.name or "")
        summary = product_summary.setdefault(
            product_name,
            {"entries": 0.0, "exits": 0.0, "current_stock": 0.0},
        )
        summary["current_stock"] += float(inventory.quantity or 0)

    for movement, product in movement_rows:
        product_name = str(product.name or "")
        summary = product_summary.setdefault(
            product_name,
            {"entries": 0.0, "exits": 0.0, "current_stock": 0.0},
        )
        qty = abs(float(movement.quantity or 0))
        if movement.movement_type in (MovementType.IN, MovementType.RETURN):
            summary["entries"] += qty
        elif movement.movement_type in (MovementType.OUT, MovementType.SALE):
            summary["exits"] += qty

    stock_rows = [["Produto", "Entradas", "Saidas", "Estoque atual"]]
    sorted_products = sorted(
        product_summary.items(),
        key=lambda item: (
            0 if item[1]["exits"] > 0 else 1,  # Produtos vendidos primeiro
            -item[1]["exits"],                 # Maior saída no topo
            item[0].lower(),
        ),
    )
    sold_row_indexes = []
    for product_name, summary in sorted_products:
        row_index = len(stock_rows)
        stock_rows.append(
            [
                product_name,
                _fmt_num(summary["entries"], 0),
                _fmt_num(summary["exits"], 0),
                _fmt_qty(summary["current_stock"]),
            ]
        )
        if summary["exits"] > 0:
            sold_row_indexes.append(row_index)
    if len(stock_rows) == 1:
        stock_rows.append(["Sem estoque controlado", "0", "0", "0"])

    story.append(Paragraph("Estoque por produto (Entradas/Saidas/Atual)", styles["Heading3"]))
    stock_table = Table(stock_rows, colWidths=[240, 90, 90, 90])
    stock_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F2F2F2")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ALIGN", (2, 1), (-1, -1), "RIGHT"),
            ]
        )
    )
    for row_index in sold_row_indexes:
        stock_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, row_index), (-1, row_index), colors.HexColor("#E8F5E9")),
                    ("TEXTCOLOR", (0, row_index), (-1, row_index), colors.HexColor("#1B5E20")),
                ]
            )
        )
    story.append(stock_table)
    story.append(Spacer(1, 12))

    movement_table_rows = [["Hora", "Produto", "Tipo", "Qtd", "Obs"]]
    for movement, product in movement_rows[:80]:
        movement_table_rows.append(
            [
                movement.created_at.strftime("%H:%M"),
                str(product.name or ""),
                str(movement.movement_type or ""),
                _fmt_num(abs(float(movement.quantity or 0))),
                str(movement.notes or movement.reference or ""),
            ]
        )
    if len(movement_table_rows) == 1:
        movement_table_rows.append(["-", "Sem movimentos hoje", "-", "0", "-"])

    story.append(Paragraph("Movimentos do dia", styles["Heading3"]))
    movement_table = Table(movement_table_rows, colWidths=[55, 170, 75, 55, 175])
    movement_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F2F2F2")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ALIGN", (3, 1), (3, -1), "RIGHT"),
            ]
        )
    )
    story.append(movement_table)

    doc.build(story)
    pdf_bytes = buffer.getvalue()
    filename = f"Stock_Dia_{start_date.strftime('%Y%m%d')}.pdf"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(io.BytesIO(pdf_bytes), media_type="application/pdf", headers=headers)


@router.get("/finance/summary.xlsx")
def get_finance_summary_excel(
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    user_id: Optional[int] = None,
    phone: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    terminal = controller.get_terminal_required(db, current_user.id)
    _require_terminal_finance_admin(db, terminal, current_user.id)
    if not start_date:
        start_date = datetime.utcnow().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    if not end_date:
        end_date = datetime.utcnow()
    filter_user_id = user_id

    summary = controller.get_financial_summary(db, terminal.id, start_date, end_date, filter_user_id)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumo Financeiro"

    ws.append(["Periodo", f"{start_date.strftime('%d/%m/%Y')} a {end_date.strftime('%d/%m/%Y')}"])
    ws.append([])
    ws.append(["Métrica", "Valor"])
    ws.append(["Entradas (vendas)", summary["gross_revenue"]])
    ws.append(["Lucro bruto", summary["gross_profit"]])
    ws.append(["Saídas (despesas)", summary["total_expenses"]])
    ws.append(["Lucro líquido", summary["net_profit"]])
    ws.append(["Nº vendas", summary["sales_count"]])
    ws.append(["Nº despesas", summary["expenses_count"]])

    ws.append([])
    ws.append(["Despesas por categoria"])
    ws.append(["Categoria", "Valor"])
    for item in summary.get("expense_breakdown") or []:
        ws.append([item.get("category_name") or "Sem categoria", item.get("total_amount") or 0])

    # Ajustar larguras
    for col in range(1, 3):
        ws.column_dimensions[get_column_letter(col)].width = 28

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    filename = f"financeiro_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}

    if phone:
        caption = f"Resumo financeiro (Excel) {start_date.strftime('%d/%m/%Y')} a {end_date.strftime('%d/%m/%Y')} (SkyPDV)."
        send_whatsapp_file(phone, filename, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", bio.getvalue(), caption=caption)
        send_whatsapp_text(phone, caption)

    return StreamingResponse(bio, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)


# ===================================================================
# FastFood stubs (compat for SkyPDV frontend)
# ===================================================================


@router.post("/fastfood/restaurants")
async def create_fastfood_restaurant(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    data = await request.form()
    name = data.get("name") or "FastFood"
    return {
        "id": 1,
        "user_id": current_user.id,
        "name": name,
        "category": data.get("category") or "restaurant",
        "is_open": False,
        "active": True,
        "phone": data.get("phone"),
        "address": data.get("address"),
    }


@router.get("/fastfood/restaurants/mine")
def list_my_restaurants(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    return []


# ===================================================================
# Accounts (Contas)
# ===================================================================

@router.get("/accounts", response_model=List[schemas.PDVAccount])
def list_accounts(
    status: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    return controller.get_accounts(db, current_user.id, status)


@router.post("/accounts", response_model=schemas.PDVAccount)
def create_account(
    data: schemas.PDVAccountCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    return controller.create_account(db, data, current_user.id)


@router.get("/accounts/{account_id}", response_model=schemas.PDVAccount)
def get_account(
    account_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    return controller.get_account(db, account_id, current_user.id)


@router.put("/accounts/{account_id}", response_model=schemas.PDVAccount)
def update_account(
    account_id: int,
    data: schemas.PDVAccountUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    return controller.update_account(db, account_id, data, current_user.id)


@router.post("/accounts/{account_id}/items", response_model=schemas.PDVAccount)
def add_account_items(
    account_id: int,
    items: List[schemas.PDVAccountItemCreate],
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    return controller.add_items_to_account(db, account_id, items, current_user.id)


@router.patch("/accounts/{account_id}/items/{item_id}", response_model=schemas.PDVAccount)
def update_account_item(
    account_id: int,
    item_id: int,
    data: schemas.PDVAccountItemUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    return controller.update_account_item(db, account_id, item_id, data, current_user.id)


@router.delete("/accounts/{account_id}/items/{item_id}", response_model=schemas.PDVAccount)
def remove_account_item(
    account_id: int,
    item_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    return controller.remove_account_item(db, account_id, item_id, current_user.id)


@router.post("/accounts/{account_id}/close", response_model=schemas.PDVAccount)
def close_account(
    account_id: int,
    data: schemas.PDVAccountClose,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    return controller.close_account(db, account_id, data, current_user.id)


@router.delete("/accounts/{account_id}")
def delete_account(
    account_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    return controller.delete_account(db, account_id, current_user.id)
